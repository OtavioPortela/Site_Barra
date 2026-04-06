from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
from django.http import HttpResponse
import logging

logger = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Cliente, OrdemServico, Servico, EstadoCabelo, TipoCabelo, CorCabelo, CorLinha
from .permissions import IsStaffOnly
from .serializers import (
    ClienteSerializer,
    OrdemServicoSerializer,
    OrdemServicoListSerializer,
    OrdemServicoStatusUpdateSerializer,
    ServicoSerializer,
    EstadoCabeloSerializer,
    TipoCabeloSerializer,
    CorCabeloSerializer,
    CorLinhaSerializer
)
from .permissions import IsOwnerOrReadOnly, CanFinalizeOS


class ClienteViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de clientes."""
    queryset = Cliente.objects.filter(ativo=True).order_by('-data_cadastro')
    serializer_class = ClienteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ativo']
    search_fields = ['nome', 'cnpj_cpf', 'email']
    ordering_fields = ['nome', 'data_cadastro']
    ordering = ['-data_cadastro']

    def perform_destroy(self, instance):
        """Soft delete - apenas marca como inativo."""
        instance.ativo = False
        instance.save()


class OrdemServicoViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de ordens de serviço."""
    queryset = OrdemServico.objects.all()
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]
    parser_classes = [MultiPartParser, FormParser, JSONParser]  # Suporte para upload de arquivos
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'cliente']
    search_fields = ['numero', 'cliente__nome', 'descricao']
    ordering_fields = ['data_criacao', 'prazo_entrega', 'valor']
    ordering = ['-data_criacao']
    pagination_class = None

    def get_serializer_class(self):
        """Retorna serializer apropriado para cada ação."""
        if self.action == 'list':
            return OrdemServicoListSerializer
        elif self.action == 'update_status':
            return OrdemServicoStatusUpdateSerializer
        return OrdemServicoSerializer

    def get_queryset(self):
        """Filtra queryset baseado em parâmetros de data."""
        queryset = super().get_queryset()

        # Filtro por faturada (só aplicar se explicitamente solicitado no histórico)
        faturada_filter = self.request.query_params.get('faturada')
        historico = self.request.query_params.get('historico')

        if historico:
            # Se for histórico, aplicar filtro de faturada se solicitado
            if faturada_filter is not None:
                faturada_bool = faturada_filter.lower() == 'true'
                queryset = queryset.filter(faturada=faturada_bool)
        else:
            # Se não for histórico (dashboard), SEMPRE excluir faturadas
            queryset = queryset.filter(faturada=False)

        # Filtro por status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filtro por data de criação
        data_inicio = self.request.query_params.get('data_inicio')
        data_fim = self.request.query_params.get('data_fim')

        if data_inicio:
            try:
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                queryset = queryset.filter(data_criacao__date__gte=data_inicio)
            except ValueError as e:
                logger.warning(f"data_inicio inválida ignorada: '{data_inicio}' — {e}")

        if data_fim:
            try:
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
                queryset = queryset.filter(data_criacao__date__lte=data_fim)
            except ValueError as e:
                logger.warning(f"data_fim inválida ignorada: '{data_fim}' — {e}")

        return queryset

    def perform_create(self, serializer):
        """Define o usuário de criação ao criar uma OS."""
        numero = serializer.validated_data.get('numero')
        if not numero:
            ultima_os = OrdemServico.objects.order_by('-id').first()
            if ultima_os and ultima_os.numero:
                try:
                    ultimo_num = int(ultima_os.numero.split('-')[-1])
                    novo_num = ultimo_num + 1
                except (ValueError, IndexError):
                    novo_num = OrdemServico.objects.count() + 1
            else:
                novo_num = 1
            numero = f"OS-{novo_num:04d}"
        serializer.save(usuario_criacao=self.request.user, numero=numero)

    def destroy(self, request, *args, **kwargs):
        """Exclusão de OS restrita a administradores."""
        if not request.user.is_staff:
            return Response(
                {'error': 'Apenas administradores podem excluir ordens de serviço.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='gerar-numero')
    def gerar_numero(self, request):
        """Endpoint para gerar o próximo número de OS."""
        ultima_os = OrdemServico.objects.order_by('-numero').first()
        if ultima_os and ultima_os.numero:
            try:
                ultimo_num = int(ultima_os.numero.split('-')[-1])
                novo_num = ultimo_num + 1
            except (ValueError, IndexError):
                novo_num = 1
        else:
            novo_num = 1

        numero = f"OS-{novo_num:04d}"
        return Response({'numero': numero}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'], url_path='atualizar-status')
    def update_status(self, request, pk=None):
        """Endpoint para atualizar apenas o status da OS."""
        ordem_servico = self.get_object()

        # Removida a restrição de finalização - agora funcionários podem finalizar
        # Apenas o patrão pode faturar (via endpoint separado)

        serializer = self.get_serializer(ordem_servico, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        novo_status = request.data.get('status')
        if novo_status == 'finalizada' and ordem_servico.status != 'finalizada':
            # Se está sendo finalizada, atualizar data_finalizacao
            ordem_servico.data_finalizacao = timezone.now()

        serializer.save()
        return Response(OrdemServicoSerializer(ordem_servico).data)

    @action(detail=True, methods=['post'], url_path='faturar', permission_classes=[IsAuthenticated])
    def faturar(self, request, pk=None):
        """Endpoint para faturar uma OS (apenas patrão)."""
        ordem_servico = self.get_object()

        # Verificar se é patrão
        if not request.user.is_staff:
            return Response(
                {'error': 'Apenas o patrão pode faturar uma ordem de serviço.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Verificar se a OS está finalizada
        if ordem_servico.status != 'finalizada':
            return Response(
                {'error': 'Apenas ordens de serviço finalizadas podem ser faturadas.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar se já foi faturada
        if ordem_servico.faturada:
            return Response(
                {'error': 'Esta ordem de serviço já foi faturada.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not ordem_servico.entregue:
            return Response(
                {'error': 'Apenas ordens de serviço entregues podem ser faturadas.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # VALIDAÇÃO: Verificar se a forma de pagamento foi definida
        if not ordem_servico.forma_pagamento:
            return Response(
                {'error': 'É necessário definir a forma de pagamento antes de faturar a ordem de serviço.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Faturar a OS
        ordem_servico.faturada = True
        ordem_servico.data_faturamento = timezone.now()
        ordem_servico.save()

        return Response(
            OrdemServicoSerializer(ordem_servico).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'], url_path='desfaturar', permission_classes=[IsAuthenticated])
    def desfaturar(self, request, pk=None):
        """Endpoint para desfaturar uma OS (devolver ao dashboard) - apenas patrão."""
        # Busca direta ignorando o filtro de faturada=False do get_queryset
        from django.shortcuts import get_object_or_404
        ordem_servico = get_object_or_404(OrdemServico, pk=pk)

        # Verificar se é patrão
        if not request.user.is_staff:
            return Response(
                {'error': 'Apenas o patrão pode desfaturar uma ordem de serviço.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Verificar se está faturada
        if not ordem_servico.faturada:
            return Response(
                {'error': 'Esta ordem de serviço não está faturada.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Desfaturar a OS (devolver ao dashboard)
        ordem_servico.faturada = False
        ordem_servico.data_faturamento = None
        ordem_servico.save()

        return Response(
            OrdemServicoSerializer(ordem_servico).data,
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['get'], url_path='exportar-excel', permission_classes=[IsAuthenticated, IsStaffOnly])
    def exportar_excel(self, request):
        """Endpoint para exportar todas as OS para Excel."""
        # Obter filtros
        status_filter = request.query_params.get('status')
        faturada_filter = request.query_params.get('faturada')
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')

        # Filtrar OS
        queryset = OrdemServico.objects.all().select_related('cliente', 'servico', 'usuario_criacao')

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if faturada_filter is not None:
            faturada_bool = faturada_filter.lower() == 'true'
            queryset = queryset.filter(faturada=faturada_bool)

        if data_inicio:
            try:
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                queryset = queryset.filter(data_criacao__date__gte=data_inicio)
            except ValueError as e:
                logger.warning(f"data_inicio inválida ignorada no export: '{data_inicio}' — {e}")

        if data_fim:
            try:
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
                queryset = queryset.filter(data_criacao__date__lte=data_fim)
            except ValueError as e:
                logger.warning(f"data_fim inválida ignorada no export: '{data_fim}' — {e}")

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordens de Serviço"

        # Cabeçalhos
        headers = [
            'Número', 'Cliente', 'Status', 'Faturada', 'Descrição',
            'Estado Cabelo', 'Tipo Cabelo', 'Cor Cabelo', 'Peso (g)', 'Tamanho (cm)',
            'Cor Linha', 'Serviço', 'Valor por Metro', 'Valor Total',
            'Data Criação', 'Prazo Entrega', 'Data Finalização', 'Data Faturamento',
            'Observações', 'Usuário Criação'
        ]

        # Estilizar cabeçalhos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Dados
        for row_num, ordem in enumerate(queryset.order_by('-data_criacao'), 2):
            ws.cell(row=row_num, column=1, value=ordem.numero)
            ws.cell(row=row_num, column=2, value=ordem.cliente.nome if ordem.cliente else '')
            ws.cell(row=row_num, column=3, value=ordem.get_status_display())
            ws.cell(row=row_num, column=4, value='Sim' if ordem.faturada else 'Não')
            ws.cell(row=row_num, column=5, value=ordem.descricao or '')
            ws.cell(row=row_num, column=6, value=ordem.get_estado_cabelo_display() if ordem.estado_cabelo else '')
            ws.cell(row=row_num, column=7, value=ordem.get_tipo_cabelo_display() if ordem.tipo_cabelo else '')
            ws.cell(row=row_num, column=8, value=ordem.cor_cabelo or '')
            ws.cell(row=row_num, column=9, value=ordem.peso_gramas or 0)
            ws.cell(row=row_num, column=10, value=ordem.tamanho_cabelo_cm or 0)
            ws.cell(row=row_num, column=11, value=ordem.cor_linha or '')
            ws.cell(row=row_num, column=12, value=ordem.servico.nome if ordem.servico else '')
            ws.cell(row=row_num, column=13, value=float(ordem.valor_metro) if ordem.valor_metro else 0)
            ws.cell(row=row_num, column=14, value=float(ordem.valor) if ordem.valor else 0)
            ws.cell(row=row_num, column=15, value=ordem.data_criacao.strftime('%d/%m/%Y %H:%M') if ordem.data_criacao else '')
            ws.cell(row=row_num, column=16, value=ordem.prazo_entrega.strftime('%d/%m/%Y') if ordem.prazo_entrega else '')
            ws.cell(row=row_num, column=17, value=ordem.data_finalizacao.strftime('%d/%m/%Y %H:%M') if ordem.data_finalizacao else '')
            ws.cell(row=row_num, column=18, value=ordem.data_faturamento.strftime('%d/%m/%Y %H:%M') if ordem.data_faturamento else '')
            ws.cell(row=row_num, column=19, value=ordem.observacoes or '')
            ws.cell(row=row_num, column=20, value=ordem.usuario_criacao.nome_completo if ordem.usuario_criacao else '')

        # Ajustar largura das colunas
        column_widths = [15, 30, 15, 10, 30, 15, 15, 15, 12, 12, 15, 20, 15, 15, 18, 15, 18, 18, 40, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        # Criar resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'ordens_servico_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class ServicoViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de serviços."""
    queryset = Servico.objects.filter(ativo=True).order_by('nome')
    serializer_class = ServicoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ativo']
    search_fields = ['nome', 'descricao']
    ordering_fields = ['nome', 'data_criacao']
    ordering = ['nome']
    pagination_class = None  # Desabilitar paginação para compatibilidade com frontend

    def perform_destroy(self, instance):
        """Soft delete - apenas marca como inativo."""
        instance.ativo = False
        instance.save()


class EstadoCabeloViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de estados do cabelo."""
    queryset = EstadoCabelo.objects.all().order_by('ordem', 'nome')
    serializer_class = EstadoCabeloSerializer
    permission_classes = [IsAuthenticated, IsStaffOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ativo']
    search_fields = ['nome', 'valor']
    ordering_fields = ['nome', 'ordem', 'data_criacao']
    ordering = ['ordem', 'nome']
    pagination_class = None


class TipoCabeloViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de tipos de cabelo."""
    queryset = TipoCabelo.objects.all().order_by('ordem', 'nome')
    serializer_class = TipoCabeloSerializer
    permission_classes = [IsAuthenticated, IsStaffOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ativo']
    search_fields = ['nome', 'valor']
    ordering_fields = ['nome', 'ordem', 'data_criacao']
    ordering = ['ordem', 'nome']
    pagination_class = None


class CorCabeloViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de cores do cabelo."""
    queryset = CorCabelo.objects.all().order_by('ordem', 'nome')
    serializer_class = CorCabeloSerializer
    permission_classes = [IsAuthenticated, IsStaffOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ativo']
    search_fields = ['nome']
    ordering_fields = ['nome', 'ordem', 'data_criacao']
    ordering = ['ordem', 'nome']
    pagination_class = None


class CorLinhaViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciamento de cores da linha."""
    queryset = CorLinha.objects.all().order_by('ordem', 'nome')
    serializer_class = CorLinhaSerializer
    permission_classes = [IsAuthenticated, IsStaffOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ativo']
    search_fields = ['nome']
    ordering_fields = ['nome', 'ordem', 'data_criacao']
    ordering = ['ordem', 'nome']
    pagination_class = None


class DebitoViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para gerenciamento de débitos de clientes parceiros."""
    serializer_class = OrdemServicoListSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['numero', 'cliente__nome', 'descricao']
    ordering_fields = ['data_criacao', 'prazo_entrega', 'valor']
    ordering = ['-data_criacao']
    pagination_class = None

    def get_queryset(self):
        """Retorna apenas OS de clientes parceiros sem forma de pagamento."""
        queryset = OrdemServico.objects.filter(
            cliente__eh_parceiro=True,
            forma_pagamento__isnull=True
        ).select_related('cliente', 'servico')

        # Filtro por parceiro_id
        parceiro_id = self.request.query_params.get('parceiro_id')
        if parceiro_id:
            queryset = queryset.filter(cliente_id=parceiro_id)

        return queryset

    @action(detail=True, methods=['patch'])
    def marcar_pago(self, request, pk=None):
        """Marca um débito como pago, atualizando a forma de pagamento."""
        debito = self.get_object()
        forma_pagamento = request.data.get('forma_pagamento')

        if not forma_pagamento:
            return Response(
                {'erro': 'Forma de pagamento é obrigatória.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if forma_pagamento not in ['dinheiro', 'pix', 'cartao_credito', 'cartao_debito']:
            return Response(
                {'erro': 'Forma de pagamento inválida.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        debito.forma_pagamento = forma_pagamento
        debito.save()

        serializer = self.get_serializer(debito)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='exportar-nota')
    def exportar_nota(self, request):
        """Exporta nota de débitos de um parceiro em formato Excel."""
        parceiro_id = request.query_params.get('parceiro_id')
        formato = request.query_params.get('formato', 'excel').lower()

        if not parceiro_id:
            return Response(
                {'erro': 'parceiro_id é obrigatório.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            parceiro = Cliente.objects.get(id=parceiro_id, eh_parceiro=True)
        except Cliente.DoesNotExist:
            return Response(
                {'erro': 'Cliente parceiro não encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Buscar débitos do parceiro
        debitos = OrdemServico.objects.filter(
            cliente_id=parceiro_id,
            forma_pagamento__isnull=True
        ).order_by('data_criacao')

        if formato == 'excel':
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Nota de Débitos"

            # Cabeçalho da empresa
            ws.merge_cells('A1:E1')
            ws['A1'] = 'BARRA CONFECCOES LTDA'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:E2')
            ws['A2'] = 'Nota de Débitos - Parceiro'
            ws['A2'].font = Font(bold=True, size=14)
            ws['A2'].alignment = Alignment(horizontal='center')

            # Dados do parceiro
            row = 4
            ws[f'A{row}'] = 'Cliente:'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = parceiro.nome
            row += 1

            if parceiro.cnpj_cpf:
                ws[f'A{row}'] = 'CNPJ/CPF:'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = parceiro.cnpj_cpf
                row += 1

            if parceiro.telefone:
                ws[f'A{row}'] = 'Telefone:'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = parceiro.telefone
                row += 1

            if parceiro.endereco:
                ws[f'A{row}'] = 'Endereço:'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = parceiro.endereco
                row += 1

            row += 1
            ws[f'A{row}'] = f'Data de Emissão: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            ws[f'A{row}'].font = Font(bold=True)

            # Cabeçalhos da tabela
            row += 2
            headers = ['Data', 'OS', 'Descrição/Serviço', 'Valor (R$)']
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Dados das OS
            total = 0
            for debito in debitos:
                row += 1
                ws.cell(row=row, column=1, value=debito.data_criacao.strftime('%d/%m/%Y') if debito.data_criacao else '')
                ws.cell(row=row, column=2, value=debito.numero)

                # Descrição ou serviço
                descricao_servico = debito.servico.nome if debito.servico else (debito.descricao or '')
                ws.cell(row=row, column=3, value=descricao_servico)

                valor = float(debito.valor) if debito.valor else 0
                ws.cell(row=row, column=4, value=valor)
                total += valor

            # Rodapé com total
            row += 2
            ws.merge_cells(f'C{row}:D{row}')
            ws[f'C{row}'] = 'TOTAL:'
            ws[f'C{row}'].font = Font(bold=True, size=12)
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            ws[f'E{row}'] = total
            ws[f'E{row}'].font = Font(bold=True, size=12)
            ws[f'E{row}'].number_format = '#,##0.00'

            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            # Criar resposta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            nome_arquivo = f'nota_debitos_{parceiro.nome.replace(" ", "_")}_{timezone.now().strftime("%Y%m%d")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

            wb.save(response)
            return response

        # PDF ainda não implementado
        return Response(
            {'erro': 'Formato PDF ainda não está disponível. Use formato "excel".'},
            status=status.HTTP_400_BAD_REQUEST
        )

